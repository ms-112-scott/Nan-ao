#!/usr/bin/env python3
"""南澳知識圖譜 — Excel -> data.json (本地版)"""
import json
import re
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
SOURCE_XLSX = REPO_ROOT / "data" / "source.xlsx"
OUT_JSON = REPO_ROOT / "data" / "data.json"
REPORT_JSON = REPO_ROOT / "data" / "_build_report.json"
PUBLIC_JSON = REPO_ROOT / "public" / "data.json"

sys.path.insert(0, str(REPO_ROOT / "scripts"))
from meta_mapping import (
    META_GROUPS, META_GROUP_COLORS, META_RELATIONS, META_RELATION_COLORS,
    map_node_group, map_link_label,
)


def cell(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def parse_year(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        y = int(v)
        if 1500 <= y <= 2200:
            return y
        return None
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r"(1[5-9]\d{2}|20\d{2}|21\d{2})", s)
    if m:
        return int(m.group(1))
    m = re.search(r"民國\s*(\d{1,3})", s)
    if m:
        return int(m.group(1)) + 1911
    return None


def split_id(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    return s if s else None


def iter_rows_dict(ws):
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return
    headers = {}
    for i, h in enumerate(header_row):
        h_clean = cell(h)
        if h_clean:
            headers[h_clean] = i
    if not headers:
        return
    for r_idx, raw in enumerate(rows, start=2):
        row = {h: cell(raw[idx]) if idx < len(raw) else None for h, idx in headers.items()}
        if all(v is None for v in row.values()):
            continue
        yield r_idx, row


def read_color_table(wb, sheet_name, key_col, value_col):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    out = {}
    for _, row in iter_rows_dict(ws):
        k = row.get(key_col)
        v = row.get(value_col)
        if k and v:
            out[str(k).strip()] = str(v).strip()
    return out


def main():
    if not SOURCE_XLSX.exists():
        print("ERROR: source.xlsx missing at " + str(SOURCE_XLSX), file=sys.stderr)
        sys.exit(1)

    print("Reading " + SOURCE_XLSX.name + " ...")
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True, read_only=True)

    nodes_color_sheet = read_color_table(wb, "nodes_color", "node_Group", "color")
    links_color_sheet = read_color_table(wb, "links_color", "link_group", "color")

    excluded = {"nodes_color", "links_color", "missing_node"}
    node_sheets = sorted([n for n in wb.sheetnames
                          if (n.startswith("nodes_") or n.startswith("node_")) and n not in excluded])
    link_sheets = sorted([n for n in wb.sheetnames
                          if (n.startswith("link_") or n.startswith("links_")) and n not in excluded])
    print("Node sheets: " + str(node_sheets))
    print("Link sheets: " + str(link_sheets))

    nodes_by_id = OrderedDict()
    info_sources = defaultdict(list)
    group_conflicts = defaultdict(list)
    unknown_groups = set()

    for sn in node_sheets:
        ws = wb[sn]
        source_label = sn.split("_", 1)[1] if "_" in sn else sn
        for row_idx, row in iter_rows_dict(ws):
            nid = split_id(row.get("id"))
            if not nid:
                continue
            grp = (row.get("node_Group") or "").strip() or "其他"
            info = (row.get("info") or "").strip()
            sy = parse_year(row.get("start_year"))
            ey = parse_year(row.get("end_year"))
            lon = row.get("Lon")
            lat = row.get("Lat") if row.get("Lat") is not None else row.get("Lat緯度")
            br = (row.get("breakthrough_note") or "").strip() if row.get("breakthrough_note") else ""

            if nid in nodes_by_id:
                existing = nodes_by_id[nid]
                if grp and existing["node_Group"] != grp:
                    group_conflicts[nid].append((sn, grp))
                if sy and (existing.get("start_year") is None or sy < existing["start_year"]):
                    existing["start_year"] = sy
                if ey and (existing.get("end_year") is None or ey > existing["end_year"]):
                    existing["end_year"] = ey
                if existing.get("Lon") is None and lon is not None:
                    existing["Lon"] = lon
                if existing.get("Lat") is None and lat is not None:
                    existing["Lat"] = lat
                if br and not existing.get("breakthrough_note"):
                    existing["breakthrough_note"] = br
            else:
                nodes_by_id[nid] = {
                    "id": nid, "node_Group": grp,
                    "start_year": sy, "end_year": ey,
                    "Lon": lon, "Lat": lat,
                    "breakthrough_note": br, "_sources": [],
                }

            if grp not in nodes_color_sheet and grp not in unknown_groups:
                if map_node_group(grp) == "其他":
                    unknown_groups.add(grp)
            if info:
                info_sources[nid].append((source_label, info))
            nodes_by_id[nid]["_sources"].append(source_label)

    final_nodes = []
    for nid, n in nodes_by_id.items():
        infos = info_sources.get(nid, [])
        if not infos:
            merged_info = ""
        elif len(infos) == 1:
            merged_info = infos[0][1]
        else:
            merged_info = " ; ".join("[" + src + "] " + txt for src, txt in infos)
        meta_group = map_node_group(n["node_Group"])
        out = {
            "id": nid, "node_Group": n["node_Group"], "meta_group": meta_group,
            "info": merged_info,
            "start_year": n["start_year"], "end_year": n["end_year"],
            "lon": n["Lon"], "lat": n["Lat"],
            "breakthrough_note": n.get("breakthrough_note") or "",
            "sources": sorted(set(n["_sources"])),
        }
        if nid in group_conflicts:
            out["_conflict"] = True
            out["_conflict_groups"] = group_conflicts[nid]
        final_nodes.append(out)

    valid_ids = set(nodes_by_id.keys())
    seen_links = set()
    final_links = []
    missing_node_refs = []
    unknown_labels = set()

    for sn in link_sheets:
        ws = wb[sn]
        for row_idx, row in iter_rows_dict(ws):
            a = split_id(row.get("Node_A"))
            b = split_id(row.get("Node_B"))
            label = (row.get("label") or "").strip() or "其他"
            info = (row.get("info") or "").strip()
            year = parse_year(row.get("Date"))
            if not a or not b:
                continue
            if a not in valid_ids:
                missing_node_refs.append({"sheet": sn, "row": row_idx, "id": a, "side": "A"})
            if b not in valid_ids:
                missing_node_refs.append({"sheet": sn, "row": row_idx, "id": b, "side": "B"})
            key = (a, b, label)
            if key in seen_links:
                continue
            seen_links.add(key)
            mr = map_link_label(label)
            if mr == "其他" and label != "其他":
                unknown_labels.add(label)
            final_links.append({
                "source": a, "target": b, "label": label,
                "meta_relation": mr, "info": info, "year": year,
            })

    valid_links = [l for l in final_links if l["source"] in valid_ids and l["target"] in valid_ids]

    sgroup = defaultdict(int)
    for n in final_nodes:
        sgroup[n["meta_group"]] += 1
    srel = defaultdict(int)
    for l in valid_links:
        srel[l["meta_relation"]] += 1

    years = [n["start_year"] for n in final_nodes if n["start_year"]]
    years += [n["end_year"] for n in final_nodes if n["end_year"]]
    year_min = min(years) if years else 1850
    year_max = max(years) if years else 2030
    breakthroughs = sum(1 for n in final_nodes if n.get("breakthrough_note"))

    data = {
        "version": datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S"),
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "stats": {
            "nodes": len(final_nodes), "links": len(valid_links),
            "links_dropped": len(final_links) - len(valid_links),
            "breakthroughs": breakthroughs,
            "by_meta_group": dict(sgroup), "by_meta_relation": dict(srel),
            "year_range": [year_min, year_max],
            "conflicts": len(group_conflicts),
            "missing_node_refs": len(missing_node_refs),
        },
        "meta_groups": [
            {"id": g, "color": META_GROUP_COLORS[g], "count": sgroup.get(g, 0)}
            for g in META_GROUPS
        ],
        "meta_relations": [
            {"id": r, "color": META_RELATION_COLORS[r], "count": srel.get(r, 0)}
            for r in META_RELATIONS
        ],
        "node_groups": sorted({n["node_Group"] for n in final_nodes}),
        "nodes": final_nodes, "links": valid_links,
    }

    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    with OUT_JSON.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    PUBLIC_JSON.parent.mkdir(parents=True, exist_ok=True)
    with PUBLIC_JSON.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))

    report = {
        "generated_at": data["generated_at"], "stats": data["stats"],
        "group_conflicts": [{"id": nid, "alternatives": alts} for nid, alts in group_conflicts.items()],
        "missing_node_refs": missing_node_refs,
        "unknown_node_groups": sorted(unknown_groups),
        "unknown_link_labels": sorted(unknown_labels),
    }
    with REPORT_JSON.open("w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    s = data["stats"]
    print()
    print("DONE")
    print("  nodes: " + str(s['nodes']) + " (conflicts " + str(s['conflicts']) + ")")
    print("  links: " + str(s['links']) + " (dropped " + str(s['links_dropped']) + ")")
    print("  breakthroughs: " + str(s['breakthroughs']))
    print("  year range: " + str(s['year_range'][0]) + "-" + str(s['year_range'][1]))
    print("  by meta_group:")
    for g in META_GROUPS:
        n = sgroup.get(g, 0)
        if n:
            print("    " + g + ": " + str(n))
    print("  by meta_relation:")
    for r in META_RELATIONS:
        n = srel.get(r, 0)
        if n:
            print("    " + r + ": " + str(n))
    if unknown_groups:
        sample = sorted(unknown_groups)[:8]
        suffix = " ..." if len(unknown_groups) > 8 else ""
        print("  WARN unmapped node_Group: " + str(sample) + suffix)
    if unknown_labels:
        sample = sorted(unknown_labels)[:8]
        suffix = " ..." if len(unknown_labels) > 8 else ""
        print("  WARN unmapped link label: " + str(sample) + suffix)
    if missing_node_refs:
        print("  WARN missing-node refs: " + str(len(missing_node_refs)))
    print()
    print("wrote " + str(OUT_JSON.relative_to(REPO_ROOT)))
    print("wrote " + str(PUBLIC_JSON.relative_to(REPO_ROOT)))
    print("wrote " + str(REPORT_JSON.relative_to(REPO_ROOT)))


if __name__ == "__main__":
    main()
